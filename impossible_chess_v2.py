# Impossible Chess Board Problem

import os
import random
import openpyxl
from openpyxl.styles import PatternFill

os.chdir(r'c:\users\tanne\documents')


# Set-up #


def create_board_binary_labels(sheet, columns):
    """Labels an 8x8 excel grid in 6-bit binary starting at 000000 top-left and ending at 111111 bot-right."""

    use_column = 0
    row = 1
    for i in range(64):  # cycles through 64 cells
        column = columns[use_column]
        row = row

        sheet[column + str(row)] = format(i, '06b')  # input sequenced 6-bit binary number

        if use_column == 7:  # stops and resets column count
            use_column = 0
            row += 1
        else:
            use_column += 1

    print('The board has been labeled in 6-bit binary.')


def create_random_heads_or_tails(sheet, columns):
    """Assigns each space (cell) with a value of 0 or 1 simulating a coin with either tails or heads facing up."""

    use_column = 0
    row = 1
    for i in range(64):
        column = columns[use_column]
        row = row

        sheet[column + str(row)] = random.randint(0, 1)  # assigns 0 or 1 randomly for each cell

        if use_column == 7:  # stops and resets column count
            use_column = 0
            row += 1
        else:               # increases column
            use_column += 1

    print('All coins have been randomly assigned.')


# Encoding, Decoding #


def decode_heads_or_tails(sheet):
    """Decodes 0s and 1s in given board."""

    all_rows = [[5, 6, 7, 8], [3, 4, 7, 8], [2, 4, 6, 8]]     # navigates rows with a 1 in place 6, place 5, place 4
    all_columns = [[5, 6, 7, 8], [3, 4, 7, 8], [2, 4, 6, 8]]  # navigates through columns

    sheet_state = []

    for rows in all_rows:  # navigates rows to count 1's
        ones_count = 0
        for row in rows:
            for column in range(1, 9):
                if sheet.cell(row=row, column=column).value == 1:
                    ones_count += 1

        if ones_count % 2 == 0:  # if even
            digit = 0
        else:  # if odd
            digit = 1

        sheet_state.append(digit)  # builds binary number backwards   

    for columns in all_columns:  # navigates columns to count 1's
        ones_count = 0
        for column in columns:
            for row in range(1, 9):
                if sheet.cell(row=row, column=column).value == 1:
                    ones_count += 1

        if ones_count % 2 == 0:  # if even
            digit = 0
        else:  # if odd
            digit = 1

        sheet_state.append(digit)  # continues to build binary number backwards

    print('The current state of the decoded board is:\n{}'.format(sheet_state))

    return sheet_state  # binary number as a list of 0's and 1's as ints


def find_cell_to_change(wb, workbook_name, key_cell, h_or_t, board):
    """Finds cell to change so board encodes to key location."""

    have = decode_heads_or_tails(h_or_t)

    change_to_int = board[key_cell].value  # changes cell name to binary location ID

    want = []
    for sym in change_to_int:
        want.append(int(sym))  # changes str values to ints

    change_cell_list = []  # compares decoded board to binary ID of key_cell
    for i in range(0, 6):
        if have[i] != want[i]:  # if digit is different
            change_cell_list.append('1')
        else:
            change_cell_list.append('0')

    change_cell_str = ''.join(change_cell_list)

    print(want, 'is the needed board state.')

    change_cell = ''
    for row in board.iter_rows(min_row=1, max_row=8, min_col=1, max_col=8):  # gets cell name of digit to change
        for cell in row:
            if str(cell.value) == change_cell_str:
                print('Changing value of cell ' + str(cell.coordinate))
                change_cell = cell.coordinate

    create_new_board = wb.copy_worksheet(h_or_t)  # creates new worksheet to save changed digit
    create_new_board.title = 'Changed Heads or Tails'
    wb.save(workbook_name)

    new_board = wb['Changed Heads or Tails']

    if new_board[change_cell].value == 1:  # changes the value of change_cell, 0 to 1 -or- 1 to 0
        new_board[change_cell] = 0
    else:
        new_board[change_cell] = 1

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    new_board[change_cell].fill = red_fill

    wb.save(workbook_name)


def find_key(wb, board):
    """Decodes new board to get key location. """

    new_board = wb['Changed Heads or Tails']
    key_list_int = decode_heads_or_tails(new_board)  # decodes changed board heads or tails values

    key_list_str = []
    for i in key_list_int:  # changes int values to str
        key_list_str.append(str(i))

    key_binary_location = ''.join(key_list_str)  # joins str list into single str

    key_location = ''
    for row in board.iter_rows(min_row=1, max_row=8, min_col=1, max_col=8):  # finds cell binary ID equal to key_cell ID
        for cell in row:
            if cell.value == key_binary_location:
                key_location = cell.coordinate

    return key_location  # returns as a str, formatted to excel column,row


# Start Commands #


def main():
    """Runs Program."""

    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']  # 8 columns for an 8x8 board

    ask_create_new = input('Do you want to create a new workbook? It is not a necessity. (Y/N): ')

    if 'Y' in ask_create_new.upper():  # creates new workbook with required data
        workbook_name = input('Enter a file name for the workbook ending in .xlsx : ')
        wb = openpyxl.Workbook()
        board = wb['Sheet']
        board.title = 'Board Labels'
        h_or_t = wb.create_sheet()
        h_or_t.title = 'Heads or Tails'

        create_board_binary_labels(board, columns)
        create_random_heads_or_tails(h_or_t, columns)

    else:  # loads workbook with required data
        workbook_name = 'impossible_chess_v1.xlsx'
        wb = openpyxl.load_workbook(workbook_name)
        board = wb['Board Labels']
        h_or_t = wb['Heads or Tails']

        create_random_heads_or_tails(h_or_t, columns)

        if 'Changed Heads or Tails' in wb.sheetnames:  # removes sheet created on last run if code has been run before
            wb.remove(wb['Changed Heads or Tails'])

    wb.save(workbook_name)

    key = input('What cell is the key in? A-H, 1-8: ').upper()

    find_cell_to_change(wb, workbook_name, key, h_or_t, board)
    print('The key is in cell: ' + find_key(wb, board))


if __name__ == '__main__':
    main()
