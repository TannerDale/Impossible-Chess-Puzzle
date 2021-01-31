# Impossible Chess Board Problem

import os
import random
import string
import openpyxl
from openpyxl.styles import PatternFill

os.chdir(r'c:\users\tanne\documents')


# Set-up #


class Chessboard:
    """Creates a simulated 8x8 chess board."""
    col_lets = list(string.ascii_uppercase[:8])
    row_nums = [num for num in range(1, 9)]
    boards = []

    def __init__(self, workbook, workbook_name, sheet, data_type):
        self.workbook = workbook
        self.wb_name = workbook_name
        self.sheet = sheet
        self.data_type = data_type
        self.data = self.create_cell_data()
        Chessboard.boards.append(self)

    def cell_labels(self):
        """Creates a list of cell labels in excel form ['A1', 'A2',...]."""
        labels = []
        for row in self.row_nums:
            for col in self.col_lets:
                labels.append(col + str(row))
        return labels

    def create_cell_data(self):
        """Creates a dict with key cells and value data of specified type."""
        cell_data = {}
        if self.data_type == 'binary':
            number = 0
            for cell in self.cell_labels():
                cell_data[cell] = (format(number, '06b'))
                number += 1
        else:
            for cell in self.cell_labels():
                cell_data[cell] = random.randint(0, 1)
        return cell_data


def fill_excel_data(coin_board):
    """Inputs cell data into the excel worksheet."""
    keys = list(coin_board.data.keys())
    for key in keys:
        coin_board.sheet[key] = coin_board.data[key]
    coin_board.workbook.save(coin_board.wb_name)


def change_bin_id_to_cell(b_binary, b_cell):
    """Gets cell id of a given binary value."""
    keys = list(b_binary.data.keys())
    values = list(b_binary.data.values())

    return keys[values.index(b_cell)]


def change_list_to_string(bin_list):
    """Changes a list of ints to a str."""
    bin_str = ''.join([str(i) for i in bin_list])

    return bin_str


def check_guess(answer, input_text):
    """Validates a user cell ID guess with the expected answer."""
    attempts = 1
    while True:
        guess = input(input_text)
        if guess.upper() != answer:
            print('That is not the correct cell.')
            if attempts == 3:
                return 'You have run out of guesses. Better luck next time.'
            else:
                attempts += 1
        else:
            return 'That is correct! Well done!'


# Encoding, Decoding #


def create_groups():
    """Creates a nested list of the groups of cells used for decoding the board. """
    group_nums = [[2, 4, 6, 8], [3, 4, 7, 8], [5, 6, 7, 8]]
    lets = Chessboard.col_lets

    column_groups = []
    for i in range(3):
        let_lst = [x for x in lets if lets.index(x) + 1 in group_nums[i]]
        full_scope_col = []
        for let in let_lst:
            full_scope_col += [let + str(i) for i in range(1, 9)]
        column_groups.append(full_scope_col)

    row_groups = []
    for i in range(3):
        full_scope_row = []
        for num in group_nums[i]:
            full_scope_row += [let + str(num) for let in lets]
        row_groups.append(full_scope_row)

    return column_groups + row_groups  # as a list


def decode_heads_or_tails(coin_board):
    """Decodes 0s and 1s in given board."""
    groups = create_groups()

    sheet_state = []
    for group in groups:
        ones = 0
        for cell in group:
            if coin_board.data[cell] == 1:
                ones += 1
        if ones % 2 == 0:
            digit = 0
        else:
            digit = 1
        sheet_state.insert(0, digit)

    sheet_state_str = change_list_to_string(sheet_state)
    print('The current state of the decoded board is:\n{}'.format(sheet_state_str))
    return sheet_state_str  # as a list


# Changing Cell #


def find_cell_to_change(bin_board, coin_board, key_cell):
    """Finds cell to change so board encodes to key location."""
    have = decode_heads_or_tails(coin_board)

    key_bin = bin_board.data[key_cell]
    print(key_bin)
    change_list = [1 if key_bin[i] != have[i] else 0 for i in range(6)]
    change_str = change_list_to_string(change_list)
    print(change_str)
    change_cell = change_bin_id_to_cell(bin_board, change_str)

    return change_cell  # as an excel ID str


def change_required_cell(coin_board, cell):
    """Changes the required cell. FIlls data highlights the changed cell in Excel."""
    if coin_board.data[cell] == 1:
        coin_board.data[cell] = 0
    else:
        coin_board.data[cell] = 1

    fill_excel_data(coin_board)
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    coin_board.sheet[cell].fill = red_fill
    coin_board.workbook.save(coin_board.wb_name)


# Finding Key #


def find_key(bin_board, coin_board):
    """Decodes new board to get key location."""
    key_list = decode_heads_or_tails(coin_board)
    key_str = change_list_to_string(key_list)

    key_location = change_bin_id_to_cell(bin_board, key_str)

    return key_location  # as a str


# Running Options #


def program_flip_and_find(bin_board, coins, coin_copy):
    """Program finds and flips the needed coin, then finds the key cell and prints the location."""
    key = input('What cell is the key in? A-H, 1-8: ').upper()
    to_change = find_cell_to_change(bin_board, coins, key)
    print('Changing cell {}'.format(to_change))

    change_required_cell(coin_copy, to_change)
    key_location = find_key(bin_board, coin_copy)

    print('The key is in cell: ' + key_location)


def program_flip_user_find(bin_board, coin_copy):
    """Program randomly places key and flips the needed coin for the user to find. Verifies players guess."""
    print('The key has been placed under a random key. I have changed the coin.')
    print('Open the worksheet Changed Heads or Tails in {} to see the current board state.'.format(bin_board.wb_name))

    key = random.choice(bin_board.cell_labels())
    change_cell = find_cell_to_change(bin_board, coin_copy, key)
    change_required_cell(coin_copy, change_cell)
    print(find_key(bin_board, coin_copy))  # uncomment to check following loop

    print(check_guess(key, 'What cell is the key in? '))


def user_flip(bin_board, coin_board, coin_copy):
    """Program randomly assigns heads or tails and places key. User decides which key to flip and finds the key."""
    key = 'A3'
    instructions = """
    I have laid out the coins and placed the key under cell {0}
    You may flip one coin to communicate the location.
    Go to {1} and change the cell in {2} to encode the board. Save change when done.
    """
    print(instructions.format(key, bin_board.wb_name, coin_copy.sheet.title))
    answer = find_cell_to_change(bin_board, coin_board, key)
    print(answer)

    print(check_guess(answer, 'What cell should you change? '))


def get_running_option(bin_board, coin_board, coin_copy):
    """Asks for and runs the users wanted running options."""
    options = ('''How do you want to run the program? Select:
    1) You input a the key location and the program changes the correct coin to flip then decodes.
    2) The program selects a key, then changes a coin for you to decode the location.
    3) The program selects a key and you flip the correct coin to encode the board?
    Choose an option: ''')
    answer = input(options)

    if '1' in answer:
        program_flip_and_find(bin_board, coin_board, coin_copy)
    elif '2' in answer:
        program_flip_user_find(bin_board, coin_copy)
    else:
        user_flip(bin_board, coin_board, coin_copy)


# Main #


def main():
    """Runs Program."""
    ask_create_new = input('Do you want to create a new workbook? It is not a necessity. (Y/N): ')
    if 'Y' in ask_create_new.upper():
        workbook_name = input('Enter a name for the workbook ending in .xlsx: ')
        wb = openpyxl.Workbook()
        wb.create_sheet().title = 'Board Labels'
        wb.create_sheet().title = 'Heads or Tails'
        wb.save(workbook_name)
    else:
        workbook_name = 'impossible_chess_v1.xlsx'
        wb = openpyxl.load_workbook(workbook_name)
        if 'Changed Heads or Tails' in wb.sheetnames:
            wb.remove(wb['Changed Heads or Tails'])
            wb.save(workbook_name)
    wb.create_sheet().title = 'Changed Heads or Tails'
    wb.save(workbook_name)

    bin_board = Chessboard(wb, workbook_name, wb['Board Labels'], 'binary')
    coin_board = Chessboard(wb, workbook_name, wb['Heads or Tails'], 'coin')
    coin_copy = Chessboard(wb, workbook_name, wb['Heads or Tails'], 'coin')
    coin_copy.__dict__.update(coin_board.__dict__)
    coin_copy.sheet = wb['Changed Heads or Tails']

    [fill_excel_data(b) for b in Chessboard.boards if b is not coin_copy]
    wb.save(workbook_name)

    get_running_option(bin_board, coin_board, coin_copy)
    wb.close()


if __name__ == '__main__':
    main()
